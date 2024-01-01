from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook
from rest_framework import viewsets, permissions

from enfermeros.forms import EnfermeroFormulario
from enfermeros.models import Enfermero, Datos, Enfermedad
from enfermeros.serializers import EnfermeroSerializer, DatosSerializer, EnfermedadSerializer

# Create your views here.

EnfermeroFormulario = modelform_factory(Enfermero, exclude=['activo', ])


def agregar_enfermero(request):
    global formulario
    pagina = loader.get_template('agregar.html')
    if request.method == 'GET':
        formulario = EnfermeroFormulario
    elif request.method == 'POST':
        formulario = EnfermeroFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))


def modificar_enfermero(request, id):
    global formulario
    pagina = loader.get_template('modificar.html')
    enfermero = get_object_or_404(Enfermero, pk=id)
    if request.method == 'GET':
        formulario = EnfermeroFormulario(instance=enfermero)
    elif request.method == 'POST':
        formulario = EnfermeroFormulario(request.POST, instance=enfermero)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))


def ver_enfermero(request, id):
    enfermero = Enfermero.objects.get(pk=id)
    enfermero = get_object_or_404(Enfermero, pk=id)
    datos = {'enfermero': enfermero}
    print(enfermero)
    pagina = loader.get_template('ver.html')
    return HttpResponse(pagina.render(datos, request))


def eliminar_enfermero(request, id):
    enfermero = get_object_or_404(Enfermero, pk=id)
    if enfermero:
        enfermero.delete()
        return redirect('inicio')


def generar_reporte(request, *args, **kwargs):
    # Obtenemos todas las personas de nuestra base de datos
    personas = Enfermero.objects.order_by('apellido', 'nombre')

    # Creamos el libro de trabajo
    wb = Workbook()

    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active

    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE ENFERMEROS'

    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')

    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'ID'
    ws['C3'] = 'CEDULA'
    ws['D3'] = 'NOMBRE'
    ws['E3'] = 'APELLIDO'
    ws['F3'] = 'ESPECIALIDAD'
    ws['G3'] = 'CORREO'
    cont = 4

    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for persona in personas:
        ws.cell(row=cont, column=2).value = persona.id
        ws.cell(row=cont, column=3).value = persona.cedula
        ws.cell(row=cont, column=4).value = persona.nombre
        ws.cell(row=cont, column=5).value = persona.apellido
        ws.cell(row=cont, column=6).value = persona.especialidad
        ws.cell(row=cont, column=7).value = persona.correo
        cont = cont + 1

    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteEnfermerosExcel.xlsx"

    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


class EnfermeroViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Enfermero.objects.all().order_by('-apellido')
    serializer_class = EnfermeroSerializer
    permission_classes = [permissions.IsAuthenticated]


class DatosViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Datos.objects.all()
    serializer_class = DatosSerializer
    permission_classes = [permissions.IsAuthenticated]


class EnfermedadViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Enfermedad.objects.all()
    serializer_class = EnfermedadSerializer
    permission_classes = [permissions.IsAuthenticated]
